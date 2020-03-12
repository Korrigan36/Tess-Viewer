# -*- coding: utf-8 -*-
"""
Created on Fri Jun 16 09:51:31 2017

@author: v-stpurc
"""
import sys
from PyQt5.QtWidgets import (QApplication, QWidget, QLabel, QRadioButton, QVBoxLayout, QCheckBox, QProgressBar,
    QGroupBox, QComboBox, QLineEdit, QPushButton, QMessageBox, QInputDialog, QDialog, QDialogButtonBox, QSlider, QGridLayout, QHBoxLayout, QFileDialog, QListWidget)
from PyQt5.QtGui import *
from PyQt5.QtCore import *
from PyQt5 import QtCore , QtGui

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.formatting.rule import Rule
from openpyxl.drawing.image import Image

import pandas as pd
import matplotlib.pyplot as plt
from matplotlib.backends.backend_qt5agg import FigureCanvasQTAgg as FigureCanvas
from matplotlib.backends.backend_qt5agg import NavigationToolbar2QT as NavigationToolbar
 
import datetime
import time
     
class MainWindow(QWidget):
    
    column_Dict = {'Board ID': '', 'Voltage': '', 'Board Temperature': '', 'Value': ''}
    row_Dict    = {'Headers': 7}
    sheetNames = []
    dataSheetIndex = []
    dataSheetNames = []
    headerRow = 0
    valueColumn = 0
    activeSheetIndex = 0


    def __init__(self):
        super(MainWindow, self).__init__()
        
        self.dataFileName = QFileDialog.getOpenFileName(self, 'Open file', 'C:\\Users\\v-stpur\\Documents\\Tess_Work',"Excel files (*.xlsx)")
        print ("Filename is: " + str(self.dataFileName[0]))
        # Open the Data file
        self.data_workbook = load_workbook(self.dataFileName[0], data_only=True)

        self.initUI()
        self.findDataSheets()
#        self.findRowsAndColumns()
        self.populateTestTypeCb()
        self.initValueCb()
        
       
    def initUI(self):
        self.setGeometry(300, 300, 600, 200)
        self.setWindowTitle('HDMI Data Visualization')
        self.setWindowIcon(QIcon('xbox_icon.ico')) 
        
        instrumentChoiceGroupBox  = QGroupBox()
        instrumentChoiceLayout    = QHBoxLayout()

        instrumentChoiceGroupBox.setLayout(instrumentChoiceLayout)

        testTypeGroupBox  = QGroupBox()
        testTypeLayout    = QHBoxLayout()
        
        self.testTypeCb = QComboBox()
        self.testTypeCb.activated[str].connect(self.onChangedTestType)
        testTypeLayout.addWidget(self.testTypeCb)
        testTypeGroupBox.setLayout(testTypeLayout)

        outputGroupBox  = QGroupBox()
        outputLayout    = QHBoxLayout()
        
        self.outputCb = QComboBox()
#        self.outputCb.activated[str].connect(self.onChangedValue)
        outputLayout.addWidget(self.outputCb)
        outputGroupBox.setLayout(outputLayout)

        startButtonGroupBox  = QGroupBox()
        startButtonLayout    = QHBoxLayout()
        self.startStopButton = QPushButton('Start Plotting', self)
        self.startStopButton.setGeometry(800, 70, 180, 50)

        self.font = QFont()
        self.font.setBold(True)
        self.font.setPointSize(16)
        self.startStopButton.setFont(self.font)
        self.startStopButton.setStyleSheet('QPushButton {background-color: #A3C1DA; color: black;}')
        self.startStopButton.setText("Start Plotting")
        self.startStopButton.clicked[bool].connect(self.plot)
        startButtonLayout.addWidget(self.startStopButton)
        startButtonGroupBox.setLayout(startButtonLayout)
 
        quitButtonGroupBox  = QGroupBox()
        quitButtonLayout    = QHBoxLayout()
        self.font.setPointSize(12)
        self.quitButton = QPushButton('Quit', self)
        self.quitButton.setFont(self.font)
        self.quitButton.setGeometry(890, 260, 100, 30)
        self.quitButton.clicked[bool].connect(self.closeEventLocal)
        quitButtonLayout.addWidget(self.quitButton)
        quitButtonGroupBox.setLayout(quitButtonLayout)
        
        grid = QGridLayout()
        grid.setColumnStretch(0,5)
        grid.setColumnStretch(1,5)
        grid.addWidget(testTypeGroupBox, 0, 0)
        grid.addWidget(outputGroupBox, 0, 1)
        grid.addWidget(startButtonGroupBox, 1, 0)
        grid.addWidget(quitButtonGroupBox, 1, 1)
        self.figure = plt.figure(figsize=(15,5))    
        self.canvas = FigureCanvas(self.figure)     
        grid.addWidget(self.canvas, 2,0,1,2)
        self.setLayout(grid)

        self.show()


        

    def closeEventLocal(self, event):
        print ("closeevent")


    def is_number(self, s):
        try:
            float(s)
            return True
        except ValueError:
            return False   
        
    def startStopTest(self):
        print ("startStopTest")
                
    def findDataSheets(self):

        print ("Searching for rows and columns")
        
        self.Num_Data_Sheets = len(self.data_workbook.sheetnames)
        print ("number of sheets in workbook" + str(self.Num_Data_Sheets))
        
        for sheetIndex in range (1, self.Num_Data_Sheets):
            tempString = self.data_workbook.sheetnames[sheetIndex]
            if tempString.find("MHz") != -1:
                self.dataSheetIndex.append(sheetIndex)
                self.dataSheetNames.append(tempString)
                
        print (self.dataSheetIndex)
        print (self.dataSheetNames)

    def findHeaderRow(self, index):

            temp_sheet = self.data_workbook.worksheets[self.dataSheetIndex[index]]
            for cellIndex in range (1, 10):
                cellObj = temp_sheet.cell(cellIndex, 1) 
                if cellObj.value != None:
                    if cellObj.value.find("Board ID") != -1:
                        self.headerRow = cellIndex
#                        print ("Header Row is: " + str(self.headerRow))
                        break

    def findTrialColumn(self, index):

            temp_sheet = self.data_workbook.worksheets[self.dataSheetIndex[index]]

            for cellIndex in range (0, 100):
                cellObj = temp_sheet.cell(self.headerRow, cellIndex + 1) 
                if cellObj.value != None:
                    if cellObj.value.find("Trial") != -1:
                        self.column_Dict['Trial'] = cellIndex
                        print ("Trial Column is: " + str(cellIndex))

    def findRowsAndColumns(self, index, valueString):

            temp_sheet = self.data_workbook.worksheets[self.dataSheetIndex[index]]

            for cellIndex in range (0, 100):
                cellObj = temp_sheet.cell(self.headerRow, cellIndex + 1) 
                if cellObj.value != None:
                    if cellObj.value.find("Board ID") != -1:
                        self.column_Dict['Board ID'] = cellIndex 
                    elif cellObj.value.find(valueString) != -1:
                        self.column_Dict['Value'] = cellIndex
                    elif cellObj.value.find("Measured V_3P3STBY") != -1:
                        self.column_Dict['Voltage'] = cellIndex
                    elif cellObj.value.find("Board Temperature") != -1:
                        self.column_Dict['Board Temperature'] = cellIndex
                        
            print (self.column_Dict)

    def onChangedTestType(self):
        print ("changed test type")  
        index = self.testTypeCb.currentIndex()  
        print (index) 
        self.populateValueCb(index) 

    def initValueCb(self):

        self.findHeaderRow(0)
        self.findTrialColumn(0)
        temp_sheet = self.data_workbook.worksheets[self.dataSheetIndex[0]]
        for cellIndex in range (self.column_Dict['Trial'], 200):
            cellObj = temp_sheet.cell(self.headerRow, cellIndex + 2) 
            if cellObj.value != None:
                self.outputCb.addItem(cellObj.value)

    def populateValueCb(self, index):

        self.findHeaderRow(index)
        self.findTrialColumn(index)
        self.outputCb.clear()
        temp_sheet = self.data_workbook.worksheets[self.dataSheetIndex[index]]
        for cellIndex in range (self.column_Dict['Trial'], 200):
            cellObj = temp_sheet.cell(self.headerRow, cellIndex + 2) 
            if cellObj.value != None:
                self.outputCb.addItem(cellObj.value)
                        
    def populateTestTypeCb(self):

        for sheetIndex in range (0, len(self.dataSheetNames)):
            self.testTypeCb.addItem(self.dataSheetNames[sheetIndex])
                        
    def onChangedValue(self):
        print ("changed test type")  
        index = self.outputCb.currentIndex()  
        self.valueColumn = index + self.column_Dict['Trial'] + 2
        print (self.valueColumn)

            
#            print (self.column_Dict)
#            
##            df = pd.read_excel(self.dataFileName[0], sheet_name='594 MHz', usecols=cols)
##            df = pd.read_excel(self.dataFileName[0], sheet_name=self.data_workbook.sheetnames[self.dataSheetIndex[sheetIndex]], usecols=cols)
#            self.df = pd.read_excel(self.dataFileName[0], sheet_name=self.data_workbook.sheetnames[self.dataSheetIndex[sheetIndex]], usecols=self.column_Dict.values(), skiprows=headerRow + 1)
#            self.df.head()
#            self.df.columns = {0:'Board ID', 1:'Voltage', 2:'Board Temperature', 3:'Trial', 4:'Board Temperature'}
#            print (self.df.columns)
#            print (self.df)
#
##            df.plot(kind='scatter',x=1,y=3,color='red')
##            plt.show()
##            self.canvas.draw()
#            self.plot()
            

    def plot(self):

        valueIndex = self.outputCb.currentIndex()  
        testTypeIndex = self.testTypeCb.currentIndex()  

        index = self.outputCb.currentIndex()  
        self.valueColumn = index + self.column_Dict['Trial'] + 2
        print (self.valueColumn)

        temp_sheet = self.data_workbook.worksheets[self.dataSheetIndex[testTypeIndex]]
        cellObj = temp_sheet.cell(self.headerRow, self.valueColumn) 
        if cellObj.value != None:
            print (cellObj.value)

            self.findRowsAndColumns(testTypeIndex, cellObj.value)

            self.df = pd.read_excel(self.dataFileName[0], sheet_name=self.data_workbook.sheetnames[self.dataSheetIndex[testTypeIndex]], usecols=self.column_Dict.values(), skiprows=self.headerRow)
#            self.df.head()
            self.df.columns = {'Board ID':'', 'Voltage':'Voltage', 'Board Temperature':'Board Temperature', 'Trial':'Trial', 'Value':'Value'}
#            self.df.rename(columns = {0:'Board ID', 1:'Voltage', 2:'Board Temperature', 3:'Trial', 4:'Value'})
#            self.df.sort_values('Board Temperature')
#            print (self.df.columns)
            print (self.df)

#        self.figure.clear()
#        ax =  self.figure.add_subplot(111) 
#        self.df.plot(kind='scatter',x=1,y=4,color='red', ax=ax)
#        self.canvas.draw()
    
if __name__ == '__main__':
    
    app = QCoreApplication.instance()
    if app is None:
        app = QApplication(sys.argv)
    ex = MainWindow()
    app.exec_()  
#    sys.exit(app.exec_())  
